import pandas as pd
import logging
import os
from datetime import datetime
from typing import Optional, List, Dict

# --- MSM CONFIGURATION ---

# Priority mapping as per conversion.md
PRIORITY_MAPPING = {
    'Not set': 'P3 (Low)',
    'Minor': 'P3 (Low)', 
    'Medium': 'P2 (Medium)',
    'Major': 'P1 (High)'
}

# MSM output column order - exact as specified
MSM_FINAL_COLUMN_ORDER: List[str] = [
    "S.No",
    "Tower",
    "Application",
    "JIRA ID",
    "Priority",
    "Issue Summary",
    "Assignee",
    "Platform / Content / Data",
    "Status",
    "Issue Status",
    "Month",
    "Issue Creation Time mm/dd/yyyy hh:mm:ss am/pm",
    "Issue Assigned Time (CTS)mm/dd/yyyy hh:mm:ss am/pm",
    "CTS Response Time mm/dd/yyyy hh:mm:ss am/pm",
    "Response SLA Met?",
    "CTS Resolution Time mm/dd/yyyy hh:mm:ss am/pm",
    "Resolution SLA Met?",
    "Last updated Date",
    "Service Category",
    "Request Type",
    "Causal Code",
    "Resolution Code",
    "High Level Debt Classification",
    "Technical Debt Classification",
    "Functional Debt Classification",
    "Operational Debt Classification",
    "Knowledge Debt Classification",
    "Time Spent()"
]

# --- LOGGING ---

def setup_msm_logger(name: str = 'ApplensTransformer') -> logging.Logger:
    logger = logging.getLogger(name)
    logger.setLevel(logging.INFO)

    # Only add file handler if not already present
    has_file_handler = any(isinstance(h, logging.FileHandler) and 'msm_conversion.log' in str(h.baseFilename) for h in logger.handlers)
    
    if not has_file_handler:
        # File Handler only - GUI handler will be added by main_gui.py
        file_handler = logging.FileHandler('msm_conversion.log', delay=True)
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
        file_handler.setFormatter(formatter)
        logger.addHandler(file_handler)

    return logger

logger = setup_msm_logger('ApplensTransformer')

# --- MSM PIPELINE FUNCTIONS ---

def load_jira_data(file_path: str) -> Optional[pd.DataFrame]:
    logger.info(f"Phase 1: Reading Jira CSV file from {file_path}")
    
    if not os.path.exists(file_path):
        logger.error(f"File not found: {file_path}")
        raise FileNotFoundError(f"Input file does not exist: {file_path}")

    try:
        # First, read only headers to identify required columns
        try:
            header_df = pd.read_csv(file_path, nrows=0, encoding='utf-8')
        except UnicodeDecodeError:
            header_df = pd.read_csv(file_path, nrows=0, encoding='latin1')
        
        all_columns = list(header_df.columns)
        logger.info(f"Found {len(all_columns)} total columns in file")
        
        # Find matching columns using precise matching to avoid duplicates
        columns_to_read = []
        column_map = {}
        found_columns = set()
        
        for col in all_columns:
            col_lower = col.lower().strip()
            
            # Issue Key - exact match first, then flexible
            if 'Issue Key' not in found_columns:
                if col_lower == 'issue key' or col_lower == 'key':
                    columns_to_read.append(col)
                    column_map[col] = 'Issue Key'
                    found_columns.add('Issue Key')
                    continue
            
            # Project Name - be more specific
            if 'Project Name' not in found_columns:
                if col_lower == 'project name' or col_lower == 'project':
                    columns_to_read.append(col)
                    column_map[col] = 'Project Name'
                    found_columns.add('Project Name')
                    continue
            
            # Summary
            if 'Summary' not in found_columns and col_lower == 'summary':
                columns_to_read.append(col)
                column_map[col] = 'Summary'
                found_columns.add('Summary')
                continue
            
            # Assignee
            if 'Assignee' not in found_columns and 'assignee' in col_lower:
                columns_to_read.append(col)
                column_map[col] = 'Assignee'
                found_columns.add('Assignee')
                continue
            
            # Priority
            if 'Priority' not in found_columns and col_lower == 'priority':
                columns_to_read.append(col)
                column_map[col] = 'Priority'
                found_columns.add('Priority')
                continue
            
            # Status
            if 'Status' not in found_columns and col_lower == 'status':
                columns_to_read.append(col)
                column_map[col] = 'Status'
                found_columns.add('Status')
                continue
            
            # Platform
            if 'Platform' not in found_columns and 'platform' in col_lower:
                columns_to_read.append(col)
                column_map[col] = 'Platform'
                found_columns.add('Platform')
                continue
            
            # Created
            if 'Created' not in found_columns and 'created' in col_lower:
                columns_to_read.append(col)
                column_map[col] = 'Created'
                found_columns.add('Created')
                continue
            
            # Updated
            if 'Updated' not in found_columns and 'updated' in col_lower:
                columns_to_read.append(col)
                column_map[col] = 'Updated'
                found_columns.add('Updated')
                continue
            
            # Resolved
            if 'Resolved' not in found_columns and 'resolved' in col_lower:
                columns_to_read.append(col)
                column_map[col] = 'Resolved'
                found_columns.add('Resolved')
                continue
            
            # Worklog
            if 'Worklog' not in found_columns and ('worklog' in col_lower or 'time spent' in col_lower):
                columns_to_read.append(col)
                column_map[col] = 'Worklog'
                found_columns.add('Worklog')
                continue
        
        logger.info(f"Reading {len(columns_to_read)} relevant columns: {list(column_map.values())}")
        
        # Read only the required columns
        try:
            df = pd.read_csv(file_path, usecols=columns_to_read, encoding='utf-8')
        except UnicodeDecodeError:
            logger.warning("UTF-8 decode failed, retrying with latin1.")
            df = pd.read_csv(file_path, usecols=columns_to_read, encoding='latin1')
        
        # Apply column mapping and handle any remaining duplicates
        df = df.rename(columns=column_map)
        
        # Remove duplicate columns if any exist
        df = df.loc[:, ~df.columns.duplicated()]
        
        logger.info(f"Successfully loaded {len(df)} rows with unique columns: {list(df.columns)}")
        return df
        
    except Exception as e:
        logger.critical(f"Failed to read CSV file: {str(e)}")
        raise e

def apply_msm_transformations(df: pd.DataFrame) -> pd.DataFrame:
    logger.info("Phase 2: Applying MSM transformations...")
    
    # Create new dataframe with MSM structure
    msm_df = pd.DataFrame()
    
    # Sequential numbering (1, 2, 3...)
    msm_df['S.No'] = range(1, len(df) + 1)
    
    # Tower - Direct copy from Project Name
    msm_df['Tower'] = df.get('Project Name', '')
    
    # Application - Constant (same for all rows)
    msm_df['Application'] = 'HMOF'
    
    # JIRA ID - Direct copy from Issue Key
    msm_df['JIRA ID'] = df.get('Issue Key', '')
    
    # Priority - Map according to conversion.md
    if 'Priority' in df.columns:
        msm_df['Priority'] = df['Priority'].map(PRIORITY_MAPPING).fillna('P3 (Low)')
    else:
        msm_df['Priority'] = 'P3 (Low)'
    
    # Issue Summary - Direct copy from Summary
    msm_df['Issue Summary'] = df.get('Summary', '')
    
    # Assignee - Direct copy
    msm_df['Assignee'] = df.get('Assignee', '')
    
    # Platform / Content / Data - Direct copy from Platform
    msm_df['Platform / Content / Data'] = df.get('Platform', '')
    
    # Status - Direct copy
    msm_df['Status'] = df.get('Status', '')
    
    # Issue Status - Direct copy (same as Status)
    msm_df['Issue Status'] = df.get('Status', '')
    
    # Month - Constant based on current month
    current_month = datetime.now().strftime('%B')
    msm_df['Month'] = current_month
    
    # Issue Creation Time - Direct copy from Created
    msm_df['Issue Creation Time mm/dd/yyyy hh:mm:ss am/pm'] = df.get('Created', '')
    
    # Issue Assigned Time - Direct copy from Created (as per conversion.md)
    msm_df['Issue Assigned Time (CTS)mm/dd/yyyy hh:mm:ss am/pm'] = df.get('Created', '')
    
    # CTS Response Time - Direct copy from Updated
    msm_df['CTS Response Time mm/dd/yyyy hh:mm:ss am/pm'] = df.get('Updated', '')
    
    # Response SLA Met - Constant: Yes
    msm_df['Response SLA Met?'] = 'Yes'
    
    # CTS Resolution Time - Direct copy from Resolved
    msm_df['CTS Resolution Time mm/dd/yyyy hh:mm:ss am/pm'] = df.get('Resolved', '')
    
    # Resolution SLA Met - If JIRA ID contains CSI → Yes, else → NA
    msm_df['Resolution SLA Met?'] = msm_df['JIRA ID'].apply(
        lambda x: 'Yes' if 'CSI' in str(x).upper() else 'NA'
    )
    
    # Last updated Date - Direct copy from Updated
    msm_df['Last updated Date'] = df.get('Updated', '')
    
    # Fields that need clarification - set as empty for now
    msm_df['Service Category'] = ''
    msm_df['Request Type'] = ''
    msm_df['Causal Code'] = ''
    msm_df['Resolution Code'] = ''
    msm_df['High Level Debt Classification'] = ''
    msm_df['Technical Debt Classification'] = ''
    msm_df['Functional Debt Classification'] = ''
    msm_df['Operational Debt Classification'] = ''
    msm_df['Knowledge Debt Classification'] = ''
    
    # Time Spent - Convert to hours format (input is in seconds)
    if 'Worklog' in df.columns:
        time_spent = pd.to_numeric(df['Worklog'], errors='coerce').fillna(0)
        # Convert from seconds to hours (divide by 3600)
        time_in_hours = (time_spent / 3600).round(2)
        msm_df['Time Spent()'] = time_in_hours
    else:
        msm_df['Time Spent()'] = 0.0
    
    logger.info(f"MSM transformation complete. Generated {len(msm_df)} rows.")
    return msm_df

def validate_msm_data(df: pd.DataFrame) -> pd.DataFrame:
    logger.info("Phase 3: Validating MSM data...")
    
    # Ensure all required columns exist in correct order
    for col in MSM_FINAL_COLUMN_ORDER:
        if col not in df.columns:
            df[col] = ''
    
    # Clean and validate data
    initial_count = len(df)
    df = df.dropna(subset=['JIRA ID'], how='all')
    if len(df) < initial_count:
        logger.warning(f"Dropped {initial_count - len(df)} rows with missing JIRA IDs.")
    
    # Fill missing values appropriately
    df = df.fillna('')
    
    logger.info("MSM validation complete.")
    return df

def save_msm_file(df: pd.DataFrame, output_path: str) -> bool:
    logger.info(f"Phase 4: Writing MSM output to {output_path}")
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        # Ensure columns are in exact order as specified
        df_final = df[MSM_FINAL_COLUMN_ORDER]
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "MSM Data"
        
        # Add data to worksheet
        for r in dataframe_to_rows(df_final, index=False, header=True):
            ws.append(r)
        
        # Style the header row
        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Set header row height to accommodate wrapped text
        ws.row_dimensions[1].height = 45
        
        # Apply header formatting
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Auto-adjust column widths with minimum width for headers
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            # Ensure minimum width of 15 for proper header display
            adjusted_width = max(min(max_length + 2, 50), 15)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add borders to all cells with data
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=len(df_final)+1, min_col=1, max_col=len(MSM_FINAL_COLUMN_ORDER)):
            for cell in row:
                cell.border = thin_border
                if cell.row > 1:  # Data rows
                    cell.alignment = Alignment(vertical="center")
        
        # Freeze the header row
        ws.freeze_panes = "A2"
        
        # Save the workbook
        wb.save(output_path)
        logger.info("SUCCESS: MSM transformation complete with enhanced formatting.")
        return True
        
    except Exception as e:
        logger.error(f"Failed to write MSM output file: {e}")
        # Fallback to basic Excel export
        try:
            df_final = df[MSM_FINAL_COLUMN_ORDER]
            df_final.to_excel(output_path, index=False)
            logger.info("SUCCESS: MSM transformation complete (basic format).")
            return True
        except Exception as e2:
            logger.error(f"Fallback export also failed: {e2}")
            return False

def run_msm_transformation_pipeline(input_path: str, output_path: str) -> bool:
    """Main pipeline function for Jira to MSM conversion"""
    try:
        df = load_jira_data(input_path)
        df = apply_msm_transformations(df)
        df = validate_msm_data(df)
        success = save_msm_file(df, output_path)
        return success
    except Exception as e:
        logger.error(f"MSM Pipeline failed: {str(e)}")
        return False

if __name__ == "__main__":
    # Direct execution for testing
    INPUT_FILE = 'jira (2).csv'
    OUTPUT_FILE = 'MSM_Upload_Output.xlsx'

    print(f"--- Starting MSM Transformation ---")
    
    if os.path.exists(INPUT_FILE):
        print(f"Processing file: {INPUT_FILE}")
        success = run_msm_transformation_pipeline(INPUT_FILE, OUTPUT_FILE)
        if success:
            print(f"\nSUCCESS! MSM output saved to: {OUTPUT_FILE}")
        else:
            print("\nFAILURE. Please check the logs.")
    else:
        print(f"\nNOTE: No input file '{INPUT_FILE}' found.")